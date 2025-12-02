"""Student management routes for roster uploads and profile syncing."""
from __future__ import annotations

import base64
import csv
import io
import re
from typing import Dict, Iterable, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from ..repository import student_portal_repository as portal_repo
from ..repository import student_management_repository as roster_repo
from ..schemas import ResponseBase, WorkType
from ..utils.dependencies import member_required
from ..utils.student_portal_security import hash_password

router = APIRouter(prefix="/student-management", tags=["Student Management"])

TEMPLATE_HEADERS = [
    "Enrollment Number",
    "First Name",
    "Last Name",
    "Std",
    "Div",
]

ENROLLMENT_MIN_LENGTH = 11
ENROLLMENT_MAX_LENGTH = 14
AUTO_PASSWORD_FALLBACK = "stud"

EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
CSV_MIME_TYPES = {"text/csv", "application/csv", "application/vnd.ms-excel"}


def _stream_excel(headers: List[str]) -> StreamingResponse:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Student Roster"
    sheet.append(headers)
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student_roster_template.xlsx"'},
    )


@router.get("/template")
async def download_template(current_user: dict = Depends(member_required(WorkType.STUDENT))):
    del current_user
    return _stream_excel(TEMPLATE_HEADERS)


def _normalize_value(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _generate_auto_password(first_name: str, enrollment_number: str) -> str:
    normalized_name = re.sub(r"[^A-Za-z]", "", first_name).lower()
    prefix = normalized_name[:4] if normalized_name else AUTO_PASSWORD_FALLBACK
    enrollment_digits = re.sub(r"[^0-9]", "", enrollment_number)
    suffix_source = enrollment_digits or enrollment_number
    return f"{prefix}{suffix_source[-4:]}"


def _decode_csv_payload(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "utf-16le", "utf-16be", "iso-8859-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="CSV file encoding is not supported. Please export as UTF-8.")


def _duplicate_row_payload(
    *,
    row_number: int | None,
    enrollment_number: str,
    first_name: str,
    last_name: Optional[str],
    std: str,
    division: Optional[str],
    reason: str,
) -> dict:
    return {
        "row_number": row_number,
        "enrollment_number": enrollment_number,
        "first_name": first_name,
        "last_name": last_name,
        "std": std,
        "division": division,
        "reason": reason,
    }


def _build_duplicate_report(rows: List[dict]) -> dict:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Duplicate Students"
    headers = ["Row Number", *TEMPLATE_HEADERS, "Reason"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("row_number"),
                row.get("enrollment_number"),
                row.get("first_name"),
                row.get("last_name"),
                row.get("std"),
                row.get("division"),
                row.get("reason"),
            ]
        )
    workbook.save(buffer)
    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return {"filename": "student_roster_duplicates.xlsx", "content": encoded}


@router.post("/upload", response_model=ResponseBase)
async def upload_student_roster(
    file: UploadFile = File(...),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="File is empty")

    duplicate_rows: List[dict] = []

    def _record_duplicate(**kwargs) -> None:
        duplicate_rows.append(_duplicate_row_payload(**kwargs))

    def _process_rows(rows: Iterable[dict]) -> List[dict]:
        entries: List[dict] = []
        seen_enrollments: set[str] = set()
        for index, row in enumerate(rows, start=2):
            enrollment_number = _normalize_value(row.get("Enrollment Number"))
            first_name = _normalize_value(row.get("First Name"))
            last_name = _normalize_value(row.get("Last Name")) or None
            std = _normalize_value(row.get("Std"))
            division = _normalize_value(row.get("Div")) or None

            if not enrollment_number:
                raise HTTPException(status_code=400, detail=f"Row {index}: Enrollment Number is required")
            if not (ENROLLMENT_MIN_LENGTH <= len(enrollment_number) <= ENROLLMENT_MAX_LENGTH):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Row {index}: Enrollment Number must be between "
                        f"{ENROLLMENT_MIN_LENGTH} and {ENROLLMENT_MAX_LENGTH} characters"
                    ),
                )
            if enrollment_number in seen_enrollments:
                _record_duplicate(
                    row_number=index,
                    enrollment_number=enrollment_number,
                    first_name=first_name,
                    last_name=last_name,
                    std=std,
                    division=division,
                    reason="Duplicate enrollment number in file",
                )
                continue
            if not first_name or not std:
                raise HTTPException(status_code=400, detail=f"Row {index}: First Name and Std are required")

            seen_enrollments.add(enrollment_number)
            entries.append(
                {
                    "row_number": index,
                    "enrollment_number": enrollment_number,
                    "first_name": first_name,
                    "last_name": last_name,
                    "std": std,
                    "division": division,
                    "auto_password": _generate_auto_password(first_name, enrollment_number),
                }
            )
        return entries

    def _validate_headers(headers: List[str]) -> None:
        normalized_headers = [header.strip() for header in headers]
        if normalized_headers != TEMPLATE_HEADERS:
            raise HTTPException(status_code=400, detail="Headers must be: " + ", ".join(TEMPLATE_HEADERS))

    is_csv_like = content_type in CSV_MIME_TYPES or filename.endswith(".csv")
    is_excel_like = content_type in EXCEL_MIME_TYPES or filename.endswith(".xlsx")

    if is_csv_like:
        text = _decode_csv_payload(raw_bytes)
        if "\x00" in text:
            is_excel_like = True
        else:
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise HTTPException(status_code=400, detail="Missing CSV headers")
            _validate_headers(reader.fieldnames)
            entries = _process_rows(reader)
            is_excel_like = False

    if is_excel_like:
        try:
            workbook = load_workbook(filename=io.BytesIO(raw_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl errors vary
            raise HTTPException(status_code=400, detail=f"Unable to read Excel file: {exc}") from exc
        sheet = workbook.active
        header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        _validate_headers(header_row)

        excel_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell in (None, "") for cell in row):
                continue
            row_dict = {header: (row[idx] if idx < len(row) else None) for idx, header in enumerate(header_row)}
            excel_rows.append(row_dict)
        if not excel_rows:
            raise HTTPException(status_code=400, detail="No rows found in Excel")
        entries = _process_rows(excel_rows)
    elif not is_csv_like:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX files are supported")

    if not entries:
        raise HTTPException(status_code=400, detail="No rows found in file")

    existing = roster_repo.fetch_existing_enrollments(
        admin_id,
        [entry["enrollment_number"] for entry in entries],
    )
    if existing:
        existing_set = set(existing)
        filtered_entries: List[dict] = []
        for entry in entries:
            if entry["enrollment_number"] in existing_set:
                _record_duplicate(
                    row_number=entry.get("row_number"),
                    enrollment_number=entry["enrollment_number"],
                    first_name=entry["first_name"],
                    last_name=entry.get("last_name"),
                    std=entry["std"],
                    division=entry.get("division"),
                    reason="Enrollment already exists",
                )
            else:
                filtered_entries.append(entry)
        entries = filtered_entries

    for entry in entries:
        entry.pop("row_number", None)

    if entries:
        roster_repo.insert_roster_entries(admin_id, entries)
        portal_repo.bulk_upsert_student_accounts(
            [
                {
                    "enrollment_number": entry["enrollment_number"],
                    "password_hash": hash_password(entry["auto_password"]),
                }
                for entry in entries
            ]
        )

    duplicate_report = _build_duplicate_report(duplicate_rows) if duplicate_rows else None

    if entries and duplicate_rows:
        message = "Uploaded with partial success. Some entries were duplicates."
    elif entries:
        message = "Student roster uploaded successfully"
    elif duplicate_rows:
        message = "No new students added because all records were duplicates."
    else:
        message = "No data processed."

    response_data: Dict[str, Any] = {
        "records_added": len(entries),
        "students": entries,
        "duplicate_count": len(duplicate_rows),
        "duplicates": duplicate_rows,
    }
    if duplicate_report:
        response_data["duplicate_report"] = duplicate_report

    return ResponseBase(status=bool(entries), message=message, data=response_data)


@router.get("/roster", response_model=ResponseBase)
async def list_student_roster(current_user: dict = Depends(member_required(WorkType.STUDENT))) -> ResponseBase:
    admin_id = current_user["admin_id"]
    students = roster_repo.fetch_roster_entries(admin_id)
    return ResponseBase(status=True, message="Student roster fetched successfully", data={"students": students})


@router.delete("/roster/{enrollment_number}", response_model=ResponseBase)
async def delete_roster_student(
    enrollment_number: str,
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    normalized = enrollment_number.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enrollment number is required")

    admin_id = current_user["admin_id"]
    deleted = roster_repo.delete_roster_entry(admin_id, normalized)
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found in roster")

    return ResponseBase(status=True, message="Student removed from roster", data={"enrollment_number": normalized})


def _clean_filter(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


@router.get("/students", response_model=ResponseBase)
async def list_student_profiles(
    class_filter: Optional[str] = Query(default=None, alias="class_filter"),
    division_filter: Optional[str] = Query(default=None, alias="division_filter"),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    students = roster_repo.fetch_student_profiles(
        admin_id,
        class_filter=_clean_filter(class_filter),
        division_filter=_clean_filter(division_filter),
    )
    return ResponseBase(status=True, message="Student profiles fetched successfully", data={"students": students})
